import os
import pandas as pd
import re
from openpyxl import load_workbook
from datetime import datetime

# 폴더 경로 설정
months = ['9월', '10월']
base_directory = 'C:/Users/user/Desktop/Won'
summary_file_path = os.path.join(base_directory, '9-10월.xlsx')

# Excel summary 파일 로드
wb = load_workbook(summary_file_path)

for month in months:
    directory = os.path.join(base_directory, month)
    sheet = wb[month]  # 시트 선택
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            filepath = os.path.join(directory, filename)
            match = re.match(r'(\d+)-(\d+)기획예산과(\(카드\))?_(\d+)건 (\d+,\d+)', filename)
            if match:
                start_date = match.group(1)
                end_date = match.group(2)
                card = match.group(3)
                count = match.group(4)
                amount = match.group(5).replace(',', '')

                for row in sheet.iter_rows(min_row=2):
                    if isinstance(row[0].value, datetime):
                        row_date = row[0].value.strftime('%Y%m%d')
                    else:
                        try:
                            row_date = datetime.strptime(row[0].value, '%Y-%m-%d').strftime('%Y%m%d')
                        except ValueError:
                            continue  # 날짜 형식이 잘못되었을 경우 무시하고 계속 진행

                    if row_date == start_date:
                        row[2].value = pd.to_datetime(end_date)  # "수납일" 설정
                        if card:
                            row[3].value = int(count)  # "카드 건수"
                            row[4].value = int(amount)  # "카드"
                        else:
                            row[5].value = int(count)  # "이체 건수"
                            row[6].value = int(amount)  # "이체"
                        break